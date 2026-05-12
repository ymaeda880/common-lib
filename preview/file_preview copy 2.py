# -*- coding: utf-8 -*-
# common_lib/preview/file_preview.py
# ============================================================
# File preview（共通プレビュー正本）
# ============================================================
# - PDF   : PNGを保存せず、表示時に指定ページだけPNG bytes化
# - Word  : LibreOfficeで preview.pdf に変換して保存し、指定ページをPNG bytes化
# - PPT   : LibreOfficeで preview.pdf に変換して保存し、指定ページをPNG bytes化
# - Text  : 40行ごとに仮想ページ表示
# - Excel : csv / tsv / xlsx を表表示
# - Image : そのまま表示
#
# 重要：
# - Inbox専用依存を持たない
# - Project専用依存を持たない
# - file_path / kind / preview_root / preview_id を受け取る
# - 呼び出し元側で file_path 解決や last_viewed 更新を行う
# ============================================================

from __future__ import annotations

# ============================================================
# imports（stdlib）
# ============================================================
import html
import math
import subprocess
from pathlib import Path
from typing import Optional, Tuple

# ============================================================
# imports（3rd party）
# ============================================================
import pandas as pd
import streamlit as st


# ============================================================
# constants
# ============================================================
TEXT_PREVIEW_LINES_PER_PAGE = 40
BASE_PDF_WIDTH = 600


# ============================================================
# helpers（kind）
# ============================================================
def normalize_kind(kind: str) -> str:
    # ------------------------------------------------------------
    # ファイル種類を小文字正規化する
    # ------------------------------------------------------------
    return str(kind or "").strip().lower()


# ============================================================
# helpers（preview id）
# ============================================================
def safe_preview_id(preview_id: str) -> str:
    # ------------------------------------------------------------
    # preview_root 配下に作るキャッシュディレクトリ名を安全化する
    # ------------------------------------------------------------
    s = str(preview_id or "").strip()

    if not s:
        return "default"

    out: list[str] = []

    for ch in s:
        if ch.isalnum() or ch in ("-", "_"):
            out.append(ch)
        else:
            out.append("_")

    return "".join(out)


# ============================================================
# helpers（page）
# ============================================================
def clamp_page_no(page_no: int, total_pages: int) -> int:
    # ------------------------------------------------------------
    # 1-based のページ番号を 1〜total_pages に丸める
    # ------------------------------------------------------------
    try:
        p = int(page_no)
    except Exception:
        p = 1

    if total_pages <= 0:
        return 1

    if p < 1:
        return 1

    if p > total_pages:
        return total_pages

    return p


# ============================================================
# helpers（session keys）
# ============================================================
def _preview_page_key(preview_id: str) -> str:
    # ------------------------------------------------------------
    # 現在ページ保存用 session_state key
    # ------------------------------------------------------------
    return f"file_preview_page__{safe_preview_id(preview_id)}"


def _preview_page_input_key(preview_id: str) -> str:
    # ------------------------------------------------------------
    # ページ番号入力 widget 用 session_state key
    # ------------------------------------------------------------
    return f"file_preview_page_input__{safe_preview_id(preview_id)}"


def _preview_zoom_key(preview_id: str) -> str:
    # ------------------------------------------------------------
    # 表示倍率 slider 用 session_state key
    # ------------------------------------------------------------
    return f"file_preview_zoom__{safe_preview_id(preview_id)}"


def _get_preview_page(preview_id: str) -> int:
    # ------------------------------------------------------------
    # session_state から現在ページを取得する
    # ------------------------------------------------------------
    try:
        return int(st.session_state.get(_preview_page_key(preview_id), 1))
    except Exception:
        return 1


def _set_preview_page(preview_id: str, page_no: int) -> None:
    # ------------------------------------------------------------
    # session_state に現在ページを保存する
    # ------------------------------------------------------------
    try:
        st.session_state[_preview_page_key(preview_id)] = max(1, int(page_no))
    except Exception:
        st.session_state[_preview_page_key(preview_id)] = 1


def _sync_preview_page_input(preview_id: str, page_no: int) -> None:
    # ------------------------------------------------------------
    # number_input 側の表示値を現在ページと同期する
    # ------------------------------------------------------------
    st.session_state[_preview_page_input_key(preview_id)] = int(max(1, page_no))


def _get_zoom_percent(preview_id: str, raw_kind: str) -> int:
    # ------------------------------------------------------------
    # 表示倍率を取得する
    # - text は初期値 70%
    # - それ以外は初期値 100%
    # ------------------------------------------------------------
    key = _preview_zoom_key(preview_id)

    if key not in st.session_state:
        if raw_kind == "text":
            st.session_state[key] = 70
        else:
            st.session_state[key] = 100

    try:
        return int(st.session_state.get(key, 100))
    except Exception:
        return 100


# ============================================================
# helpers（fitz）
# ============================================================
def try_import_fitz():
    # ------------------------------------------------------------
    # PyMuPDF(fitz) を安全に import する
    # ------------------------------------------------------------
    try:
        import fitz  # type: ignore
        return fitz
    except Exception:
        return None


# ============================================================
# helpers（PDF page count）
# ============================================================
def get_pdf_page_count(pdf_path: Path) -> Optional[int]:
    # ------------------------------------------------------------
    # PDFファイルの総ページ数を取得する
    # ------------------------------------------------------------
    fitz = try_import_fitz()

    if fitz is None:
        return None

    try:
        doc = fitz.open(str(pdf_path))
        return int(doc.page_count)
    except Exception:
        return None


# ============================================================
# helpers（PDF bytes -> PNG bytes）
# ============================================================
def pdf_page_png(
    pdf_bytes: bytes,
    *,
    page_no: int,
    max_width: int = 1200,
) -> Optional[bytes]:
    # ------------------------------------------------------------
    # PDF bytes の指定ページを PNG bytes に変換する
    # - page_no は 1-based
    # - PNGファイルは保存しない
    # ------------------------------------------------------------
    fitz = try_import_fitz()

    if fitz is None:
        return None

    try:
        doc = fitz.open(stream=pdf_bytes, filetype="pdf")
        total_pages = int(doc.page_count)

        if total_pages <= 0:
            return None

        page_no = clamp_page_no(page_no, total_pages)
        page = doc.load_page(page_no - 1)

        pix0 = page.get_pixmap()

        if pix0.width <= 0:
            return None

        zoom = float(max_width) / float(pix0.width)

        if zoom <= 0:
            zoom = 1.0

        mat = fitz.Matrix(zoom, zoom)
        pix = page.get_pixmap(matrix=mat)

        return pix.tobytes("png")

    except Exception:
        return None


# ============================================================
# public（UIなしPDF描画）
# ============================================================
def render_pdf_page_only(
    *,
    file_path: Path,
    page_no: int,
    max_width: int,
) -> Optional[int]:
    # ------------------------------------------------------------
    # UIなしでPDF指定ページを表示する
    # - ページ数を返す
    # - 表示は行う（st.image）
    # ------------------------------------------------------------
    total_pages = get_pdf_page_count(file_path)

    if total_pages is None or total_pages <= 0:
        return None

    current_page = clamp_page_no(page_no, total_pages)

    png = pdf_page_png(
        file_path.read_bytes(),
        page_no=current_page,
        max_width=max_width,
    )

    if png:
        st.image(
            png,
            width=max_width,
        )
        return total_pages

    return None



# ============================================================
# helpers（Office -> preview.pdf）
# ============================================================
def ensure_office_preview_pdf(src_path: Path, out_dir: Path) -> Optional[Path]:
    # ------------------------------------------------------------
    # Office系ファイルを LibreOffice で PDF 化する
    # - out_dir/preview.pdf を正本キャッシュにする
    # - 既に preview.pdf があれば再利用する
    # ------------------------------------------------------------
    out_dir.mkdir(parents=True, exist_ok=True)

    out_pdf = out_dir / "preview.pdf"

    if out_pdf.exists():
        return out_pdf

    try:
        r = subprocess.run(
            ["soffice", "--version"],
            capture_output=True,
            text=True,
            check=False,
        )

        if r.returncode != 0:
            return None

    except Exception:
        return None

    try:
        subprocess.run(
            [
                "soffice",
                "--headless",
                "--convert-to",
                "pdf",
                str(src_path),
                "--outdir",
                str(out_dir),
            ],
            check=True,
            capture_output=True,
            text=True,
        )

        cand = out_dir / f"{src_path.stem}.pdf"

        if cand.exists():
            cand.replace(out_pdf)

        return out_pdf if out_pdf.exists() else None

    except Exception:
        return None


def ensure_word_preview_pdf(docx_path: Path, out_dir: Path) -> Optional[Path]:
    # ------------------------------------------------------------
    # Wordファイルを PDF 化する
    # - 実体は ensure_office_preview_pdf() に集約する
    # ------------------------------------------------------------
    return ensure_office_preview_pdf(docx_path, out_dir)


# ============================================================
# helpers（xlsx）
# ============================================================
def load_xlsx_preview_df(
    xlsx_path: Path,
    *,
    sheet_name: Optional[str] = None,
    max_rows: int = 50,
    max_cols: int = 11,
) -> Tuple[list[str], Optional[str], Optional[pd.DataFrame]]:
    # ------------------------------------------------------------
    # xlsx のシート一覧を取得し、指定シートを DataFrame 化する
    # ------------------------------------------------------------
    try:
        import openpyxl
    except Exception:
        return [], None, None

    try:
        wb = openpyxl.load_workbook(
            str(xlsx_path),
            data_only=True,
            read_only=True,
        )

        sheet_names = list(wb.sheetnames)

        if not sheet_names:
            return [], None, None

        if sheet_name not in sheet_names:
            sheet_name = sheet_names[0]

        ws = wb[sheet_name]

        rows = []

        for row in ws.iter_rows(
            min_row=1,
            max_row=max_rows,
            min_col=1,
            max_col=max_cols,
            values_only=True,
        ):
            rows.append(list(row))

        df = pd.DataFrame(rows)

        try:
            df = df.astype(str)
        except Exception:
            df = df.map(lambda x: "" if x is None else str(x))

        return sheet_names, sheet_name, df

    except Exception:
        return [], None, None

# ============================================================
# helpers（text）
# ============================================================
def read_text_preview(file_path: Path) -> Tuple[Optional[str], int]:
    # ------------------------------------------------------------
    # テキストファイルをUTF-8として読み込む
    # - errors="replace" で読み込み失敗をできるだけ回避する
    # ------------------------------------------------------------
    try:
        txt = file_path.read_text(encoding="utf-8", errors="replace")
    except Exception:
        return None, 0

    return txt, len(txt)


def paginate_text_lines(
    txt: str,
    *,
    lines_per_page: int = TEXT_PREVIEW_LINES_PER_PAGE,
) -> Tuple[list[str], int]:
    # ------------------------------------------------------------
    # テキストを行単位で仮想ページ化する
    # ------------------------------------------------------------
    lines = txt.splitlines()

    if not lines:
        lines = [""]

    if lines_per_page <= 0:
        lines_per_page = TEXT_PREVIEW_LINES_PER_PAGE

    total_pages = max(1, int(math.ceil(len(lines) / lines_per_page)))

    return lines, total_pages


def slice_text_page(
    lines: list[str],
    *,
    page_no: int,
    lines_per_page: int = TEXT_PREVIEW_LINES_PER_PAGE,
) -> Tuple[str, int, int]:
    # ------------------------------------------------------------
    # 指定ページ分のテキストと行番号範囲を返す
    # ------------------------------------------------------------
    if lines_per_page <= 0:
        lines_per_page = TEXT_PREVIEW_LINES_PER_PAGE

    if not lines:
        return "", 1, 1

    start_idx = max(0, (int(page_no) - 1) * lines_per_page)
    end_idx = min(len(lines), start_idx + lines_per_page)

    page_lines = lines[start_idx:end_idx]
    page_text = "\n".join(page_lines)

    start_line_no = start_idx + 1
    end_line_no = end_idx

    return page_text, start_line_no, end_line_no


# ============================================================
# helpers（Office preview pdf resolve）
# ============================================================
def _resolve_preview_pdf_for_office(
    *,
    raw_kind: str,
    file_path: Path,
    out_dir: Path,
) -> Optional[Path]:
    # ------------------------------------------------------------
    # Word / PPT を preview.pdf に変換し、そのPathを返す
    # - 既存 preview.pdf があれば再利用する
    # - 初回変換時のみ警告と spinner を出す
    # ------------------------------------------------------------
    preview_pdf = out_dir / "preview.pdf"

    if preview_pdf.exists():
        return preview_pdf

    status_box = st.empty()

    if raw_kind == "word":
        status_box.warning("📄 Word を PDF に変換しています（初回は時間がかかります）")
    elif raw_kind == "ppt":
        status_box.warning("📊 PowerPoint を PDF に変換しています（初回は時間がかかります）")

    with st.spinner("LibreOffice で変換中…"):
        out = ensure_office_preview_pdf(file_path, out_dir)

    status_box.empty()

    if out is not None and out.exists():
        return out

    return preview_pdf if preview_pdf.exists() else None


# ============================================================
# helpers（total pages）
# ============================================================
def _get_total_pages_for_preview(
    *,
    raw_kind: str,
    file_path: Path,
    preview_root: Path,
    preview_id: str,
) -> Optional[int]:
    # ------------------------------------------------------------
    # kind別に総ページ数を取得する
    # - pdf  : PDFの総ページ数
    # - word : preview.pdf に変換後、その総ページ数
    # - ppt  : preview.pdf に変換後、その総ページ数
    # - text : 40行ごとの仮想ページ数
    # ------------------------------------------------------------
    safe_id = safe_preview_id(preview_id)

    if raw_kind == "pdf":
        return get_pdf_page_count(file_path)

    if raw_kind == "word":
        out_dir = preview_root / safe_id
        preview_pdf = _resolve_preview_pdf_for_office(
            raw_kind=raw_kind,
            file_path=file_path,
            out_dir=out_dir,
        )

        if preview_pdf is None or not preview_pdf.exists():
            return None

        return get_pdf_page_count(preview_pdf)

    if raw_kind == "ppt":
        out_dir = preview_root / safe_id
        preview_pdf = _resolve_preview_pdf_for_office(
            raw_kind=raw_kind,
            file_path=file_path,
            out_dir=out_dir,
        )

        if preview_pdf is None or not preview_pdf.exists():
            return None

        return get_pdf_page_count(preview_pdf)

    if raw_kind == "text":
        txt, _original_len = read_text_preview(file_path)

        if txt is None:
            return 1

        _lines, total_pages = paginate_text_lines(
            txt,
            lines_per_page=TEXT_PREVIEW_LINES_PER_PAGE,
        )

        return total_pages

    return None


# ============================================================
# UI（zoom）
# ============================================================
def _render_zoom_control(
    *,
    preview_id: str,
    raw_kind: str,
) -> int:
    # ------------------------------------------------------------
    # 表示倍率 slider を描画し、倍率を返す
    # ------------------------------------------------------------
    zoom_key = _preview_zoom_key(preview_id)

    if zoom_key not in st.session_state:
        if raw_kind == "text":
            st.session_state[zoom_key] = 70
        else:
            st.session_state[zoom_key] = 100

    st.slider(
        "表示倍率（%）",
        min_value=50,
        max_value=200,
        step=10,
        key=zoom_key,
    )

    zoom_percent = _get_zoom_percent(preview_id, raw_kind)

    st.caption(f"現在倍率: {zoom_percent}%")

    return zoom_percent


# ============================================================
# UI（page navigation）
# ============================================================
def _render_page_navigation(
    *,
    preview_id: str,
    total_pages: int,
) -> int:
    # ------------------------------------------------------------
    # ページ移動UIを描画し、現在ページを返す
    # ------------------------------------------------------------
    preview_page = _get_preview_page(preview_id)
    preview_page = clamp_page_no(preview_page, total_pages)

    _set_preview_page(preview_id, preview_page)

    input_key = _preview_page_input_key(preview_id)

    if input_key not in st.session_state:
        st.session_state[input_key] = int(preview_page)
    else:
        try:
            st.session_state[input_key] = int(st.session_state[input_key])
        except Exception:
            st.session_state[input_key] = int(preview_page)

    st.markdown(
        """
        <div style="height: 12px;"></div>
        <hr style="margin: 8px 0 16px 0; border: none; border-top: 1px solid #ddd;">
        """,
        unsafe_allow_html=True,
    )

    c1, c2, c3, c4, c5 = st.columns([1, 1, 1, 1, 1])

    with c1:
        if st.button("⬅ 前へ", key=f"file_preview_prev_page_{safe_preview_id(preview_id)}"):
            if preview_page > 1:
                preview_page = int(preview_page) - 1
                _set_preview_page(preview_id, preview_page)
                _sync_preview_page_input(preview_id, preview_page)

    with c2:
        if st.button("次へ ➡", key=f"file_preview_next_page_{safe_preview_id(preview_id)}"):
            if preview_page < int(total_pages):
                preview_page = int(preview_page) + 1
                _set_preview_page(preview_id, preview_page)
                _sync_preview_page_input(preview_id, preview_page)

    with c3:
        page_1based = st.number_input(
            "ページ番号",
            min_value=1,
            max_value=max(1, int(total_pages)),
            step=1,
            key=input_key,
            label_visibility="collapsed",
        )

        if int(page_1based) != int(preview_page):
            preview_page = int(page_1based)
            _set_preview_page(preview_id, preview_page)

    with c5:
        st.write(f"Page: {int(preview_page)} / {int(total_pages)}")

    return int(preview_page)


# ============================================================
# render（image）
# ============================================================
def _render_image_preview(
    *,
    file_path: Path,
    display_name: str,
) -> None:
    # ------------------------------------------------------------
    # 画像をそのまま表示する
    # ------------------------------------------------------------
    try:
        st.image(file_path.read_bytes(), caption=display_name)
    except Exception as e:
        st.error(f"画像の表示に失敗しました: {e}")


# ============================================================
# render（pdf）
# ============================================================
def _render_pdf_preview(
    *,
    file_path: Path,
    display_name: str,
    page_no: int,
    pdf_max_width: int,
) -> Optional[int]:
    # ------------------------------------------------------------
    # PDFの指定ページをPNG bytes化して表示する
    # ------------------------------------------------------------
    total_pages = get_pdf_page_count(file_path)

    if total_pages is None or total_pages <= 0:
        st.info("PDFプレビューには PyMuPDF(fitz) が必要です。")
        return None

    current_page = clamp_page_no(page_no, total_pages)

    png = pdf_page_png(
        file_path.read_bytes(),
        page_no=current_page,
        max_width=pdf_max_width,
    )

    if png:
        st.image(
            png,
            caption=f"{display_name}（PDF {current_page} / {total_pages} ページ）",
            width=pdf_max_width,
        )
        return total_pages

    st.info("PDFプレビューには PyMuPDF(fitz) が必要です。")
    return None


# ============================================================
# render（word / ppt）
# ============================================================
def _render_office_preview(
    *,
    raw_kind: str,
    file_path: Path,
    display_name: str,
    out_dir: Path,
    page_no: int,
    pdf_max_width: int,
) -> Optional[int]:
    # ------------------------------------------------------------
    # Word / PPT を preview.pdf 経由で指定ページ表示する
    # ------------------------------------------------------------
    preview_pdf = _resolve_preview_pdf_for_office(
        raw_kind=raw_kind,
        file_path=file_path,
        out_dir=out_dir,
    )

    if preview_pdf is None or not preview_pdf.exists():
        if raw_kind == "word":
            st.error("Word → PDF 変換に失敗しました。")
        elif raw_kind == "ppt":
            st.error("PowerPoint → PDF 変換に失敗しました。")
        else:
            st.error("Office → PDF 変換に失敗しました。")
        return None

    total_pages = get_pdf_page_count(preview_pdf)

    if total_pages is None or total_pages <= 0:
        st.info("PDFプレビューには PyMuPDF(fitz) が必要です。")
        return None

    current_page = clamp_page_no(page_no, total_pages)

    png = pdf_page_png(
        preview_pdf.read_bytes(),
        page_no=current_page,
        max_width=pdf_max_width,
    )

    if not png:
        st.info("PDFプレビューには PyMuPDF(fitz) が必要です。")
        return None

    if raw_kind == "word":
        label = "Word → PDF変換後"
    elif raw_kind == "ppt":
        label = "PowerPoint → PDF変換後"
    else:
        label = "Office → PDF変換後"

    st.image(
        png,
        caption=f"{display_name}（{label} {current_page} / {total_pages} ページ）",
        width=pdf_max_width,
    )

    return total_pages


# ============================================================
# render（text）
# ============================================================
def _render_text_preview(
    *,
    file_path: Path,
    display_name: str,
    page_no: int,
    text_zoom_percent: int,
) -> Optional[int]:
    # ------------------------------------------------------------
    # Textを40行ごとに仮想ページ表示する
    # ------------------------------------------------------------
    txt, original_len = read_text_preview(file_path)

    if txt is None:
        st.error("テキストの読み込みに失敗しました。")
        return None

    lines, total_pages = paginate_text_lines(
        txt,
        lines_per_page=TEXT_PREVIEW_LINES_PER_PAGE,
    )

    current_page = clamp_page_no(page_no, total_pages)

    page_text, start_line_no, end_line_no = slice_text_page(
        lines,
        page_no=current_page,
        lines_per_page=TEXT_PREVIEW_LINES_PER_PAGE,
    )

    st.caption(
        f"{display_name}（Text {current_page} / {total_pages} ページ、"
        f"行 {start_line_no}–{end_line_no}、全 {len(lines)} 行、{original_len} 文字）"
    )

    try:
        zoom_i = int(text_zoom_percent)
    except Exception:
        zoom_i = 100

    font_size_px = max(6, int(zoom_i / 5))

    safe_text = html.escape(page_text.lstrip("\r\n"))

    html_content = (
        '<html>'
        '<body style="margin:0; padding:8px;">'
        f'<pre style="margin:0; font-size:{font_size_px}px; line-height:1.25; '
        'font-family:monospace; white-space:pre-wrap; overflow-x:auto;">'
        f'{safe_text}'
        '</pre>'
        '</body>'
        '</html>'
    )

    st.components.v1.html(
        html_content,
        height=800,
        scrolling=True,
    )

    return total_pages


# ============================================================
# render（excel）
# ============================================================
def _render_excel_preview(
    *,
    file_path: Path,
) -> None:
    # ------------------------------------------------------------
    # Excel / CSV / TSV を表形式で表示する
    # ------------------------------------------------------------
    suffix = file_path.suffix.lower()

    if suffix == ".xls":
        st.info("このExcel形式（.xls）は現在プレビュー非対応です（保存・ダウンロードは可能）。")
        return

    if suffix in (".csv", ".tsv"):
        try:
            if suffix == ".tsv":
                df_prev = pd.read_csv(file_path, dtype=str, nrows=200, sep="\t")
            else:
                df_prev = pd.read_csv(file_path, dtype=str, nrows=200)
        except Exception as e:
            st.error(f"CSV/TSV の読み込みに失敗しました: {e}")
            return

        st.caption(f"{suffix.upper()}（先頭 {min(len(df_prev), 200)} 行）")
        st.dataframe(df_prev, hide_index=True)
        return


    # ------------------------------------------------------------
    # シート一覧取得
    # ------------------------------------------------------------
    sheet_names, default_sheet_name, _ = load_xlsx_preview_df(
        file_path,
        sheet_name=None,
        max_rows=1,
        max_cols=1,
    )

    if not sheet_names or default_sheet_name is None:
        st.info("Excelプレビューには openpyxl が必要です（または読み込みに失敗しました）。")
        return

    # ------------------------------------------------------------
    # シート選択UI
    # ------------------------------------------------------------
    sheet_key = f"excel_preview_sheet__{safe_preview_id(str(file_path))}"

    selected_sheet = st.selectbox(
        "シート選択",
        options=sheet_names,
        index=sheet_names.index(default_sheet_name),
        key=sheet_key,
    )

    # ------------------------------------------------------------
    # 選択シート読み込み
    # ------------------------------------------------------------
    _, sheet_name, df_prev = load_xlsx_preview_df(
        file_path,
        sheet_name=selected_sheet,
        max_rows=50,
        max_cols=11,
    )

    if sheet_name is None or df_prev is None:
        st.info("Excelシートの読み込みに失敗しました。")
        return

    st.caption(f"シート: {sheet_name}（先頭 {min(len(df_prev), 50)} 行 × 最大 11 列）")
    st.dataframe(df_prev, hide_index=True)




# ============================================================
# public（main）
# ============================================================
def render_file_preview_with_controls(
    *,
    file_path: Path,
    kind: str,
    preview_root: Path,
    preview_id: str,
    original_name: str = "",
    title: str = "④ プレビュー",
) -> None:
    # ------------------------------------------------------------
    # 汎用ファイルプレビュー本体
    #
    # 呼び出し元が行うこと：
    # - file_path を解決する
    # - kind を渡す
    # - preview_root を渡す
    # - preview_id を渡す
    # - 必要なら閲覧履歴などを更新する
    #
    # この関数が行うこと：
    # - 見出し表示
    # - image / excel / other / pdf / word / ppt / text の分岐
    # - zoom UI
    # - page navigation UI
    # - selected page rendering
    # ------------------------------------------------------------
    raw_kind = normalize_kind(kind)
    file_path = Path(file_path)
    preview_root = Path(preview_root)
    preview_id = safe_preview_id(preview_id)
    display_name = str(original_name or file_path.name)

    st.divider()
    st.subheader(title)

    if not file_path.exists():
        st.error("プレビュー対象ファイルが存在しません。")
        return

    # ------------------------------------------------------------
    # image / excel / other はページ操作なし
    # ------------------------------------------------------------
    if raw_kind == "image":
        _render_image_preview(
            file_path=file_path,
            display_name=display_name,
        )
        return

    if raw_kind == "excel":
        _render_excel_preview(file_path=file_path)
        return

    if raw_kind not in ("pdf", "word", "ppt", "text"):
        st.info(f"未対応形式です: {raw_kind or 'other'}")
        return

    # ------------------------------------------------------------
    # zoom
    # ------------------------------------------------------------
    zoom_percent = _render_zoom_control(
        preview_id=preview_id,
        raw_kind=raw_kind,
    )

    pdf_max_width = int(BASE_PDF_WIDTH * zoom_percent / 100)

    # ------------------------------------------------------------
    # total pages
    # ------------------------------------------------------------
    total_pages = _get_total_pages_for_preview(
        raw_kind=raw_kind,
        file_path=file_path,
        preview_root=preview_root,
        preview_id=preview_id,
    )

    if total_pages is None or int(total_pages) <= 0:
        if raw_kind == "word":
            st.error("Word → PDF 変換に失敗しました。")
        elif raw_kind == "ppt":
            st.error("PowerPoint → PDF 変換に失敗しました。")
        elif raw_kind == "pdf":
            st.info("PDFプレビューには PyMuPDF(fitz) が必要です。")
        elif raw_kind == "text":
            st.error("テキストの読み込みに失敗しました。")
        return

    # ------------------------------------------------------------
    # page navigation
    # ------------------------------------------------------------
    page_no = _render_page_navigation(
        preview_id=preview_id,
        total_pages=int(total_pages),
    )

    # ------------------------------------------------------------
    # render selected kind
    # ------------------------------------------------------------
    shown_total_pages: Optional[int] = None
    safe_id = safe_preview_id(preview_id)

    if raw_kind == "pdf":
        shown_total_pages = _render_pdf_preview(
            file_path=file_path,
            display_name=display_name,
            page_no=page_no,
            pdf_max_width=pdf_max_width,
        )

    elif raw_kind == "word":
        shown_total_pages = _render_office_preview(
            raw_kind=raw_kind,
            file_path=file_path,
            display_name=display_name,
            out_dir=preview_root / safe_id,
            page_no=page_no,
            pdf_max_width=pdf_max_width,
        )

    elif raw_kind == "ppt":
        shown_total_pages = _render_office_preview(
            raw_kind=raw_kind,
            file_path=file_path,
            display_name=display_name,
            out_dir=preview_root / safe_id,
            page_no=page_no,
            pdf_max_width=pdf_max_width,
        )

    elif raw_kind == "text":
        shown_total_pages = _render_text_preview(
            file_path=file_path,
            display_name=display_name,
            page_no=page_no,
            text_zoom_percent=zoom_percent,
        )

    # ------------------------------------------------------------
    # final page caption
    # ------------------------------------------------------------
    if shown_total_pages is not None:
        current_page = clamp_page_no(page_no, int(shown_total_pages))
        st.caption(f"ページ {current_page} / {int(shown_total_pages)}")